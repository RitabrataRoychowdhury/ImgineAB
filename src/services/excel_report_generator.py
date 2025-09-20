"""
Excel Report Generator for enhanced Q&A capabilities.
Generates structured Excel reports from document analysis and chat interactions.
"""

import os
import uuid
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta
import logging
import pandas as pd

from src.utils.error_handling import (
    ExcelGenerationError, handle_errors, alternative_formats, 
    ErrorType, graceful_degradation
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

from src.models.conversational import ExcelReport, ExcelSheet, ReportTemplate
from src.models.document import Document
from src.storage.document_storage import DocumentStorage


class ExcelReportGenerator:
    """
    Generates structured Excel reports with multiple sheets and formatting.
    """
    
    def __init__(self, document_storage: DocumentStorage, reports_dir: str = "data/reports"):
        self.document_storage = document_storage
        self.reports_dir = reports_dir
        self.logger = logging.getLogger(__name__)
        
        # Ensure reports directory exists
        os.makedirs(reports_dir, exist_ok=True)
        
        # Default formatting styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.data_alignment = Alignment(wrap_text=True, vertical="top")

    @handle_errors(ErrorType.EXCEL_GENERATION_ERROR)
    def generate_document_report(self, document_id: str, report_type: str = "comprehensive") -> ExcelReport:
        """
        Generate Excel report for a single document analysis.
        """
        try:
            # Get document and analysis data
            document = self.document_storage.get_document(document_id)
            if not document:
                raise ValueError(f"Document {document_id} not found")
            
            # Get analysis data (this would come from enhanced analyzer)
            analysis_data = self._extract_document_analysis_data(document_id)
            
            # Create report structure
            sheets = []
            
            if report_type == "comprehensive":
                sheets = [
                    self._create_summary_sheet(document, analysis_data),
                    self._create_risks_sheet(analysis_data.get('risks', [])),
                    self._create_commitments_sheet(analysis_data.get('commitments', [])),
                    self._create_dates_sheet(analysis_data.get('deliverable_dates', [])),
                    self._create_terms_sheet(analysis_data.get('key_terms', []))
                ]
            elif report_type == "risks_only":
                sheets = [
                    self._create_risks_sheet(analysis_data.get('risks', []))
                ]
            elif report_type == "commitments_only":
                sheets = [
                    self._create_commitments_sheet(analysis_data.get('commitments', []))
                ]
            
            # Generate Excel file
            report_id = str(uuid.uuid4())
            filename = f"document_report_{document_id}_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(self.reports_dir, filename)
            
            self._create_excel_file(sheets, file_path)
            
            # Create download URL (this would be handled by web framework)
            download_url = f"/api/reports/download/{report_id}"
            
            return ExcelReport(
                report_id=report_id,
                filename=filename,
                file_path=file_path,
                download_url=download_url,
                sheets=sheets,
                created_at=datetime.now(),
                expires_at=datetime.now() + timedelta(days=7)
            )
            
        except Exception as e:
            self.logger.error(f"Error generating document report: {str(e)}")
            # Try to generate alternative format
            try:
                return self._generate_fallback_report(document_id, analysis_data, report_type)
            except Exception as fallback_error:
                self.logger.error(f"Fallback report generation also failed: {str(fallback_error)}")
                raise ExcelGenerationError(
                    f"Failed to generate report for document {document_id}",
                    {"document_id": document_id, "report_type": report_type},
                    e
                )

    @handle_errors(ErrorType.EXCEL_GENERATION_ERROR)
    def generate_conversation_report(self, session_id: str) -> ExcelReport:
        """
        Generate Excel report from conversation history.
        """
        try:
            # Get conversation data (this would come from conversation context)
            conversation_data = self._extract_conversation_data(session_id)
            
            # Create sheets
            sheets = [
                self._create_conversation_summary_sheet(conversation_data),
                self._create_qa_history_sheet(conversation_data.get('turns', [])),
                self._create_topics_sheet(conversation_data.get('topics', []))
            ]
            
            # Generate Excel file
            report_id = str(uuid.uuid4())
            filename = f"conversation_report_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(self.reports_dir, filename)
            
            self._create_excel_file(sheets, file_path)
            
            download_url = f"/api/reports/download/{report_id}"
            
            return ExcelReport(
                report_id=report_id,
                filename=filename,
                file_path=file_path,
                download_url=download_url,
                sheets=sheets,
                created_at=datetime.now(),
                expires_at=datetime.now() + timedelta(days=7)
            )
            
        except Exception as e:
            self.logger.error(f"Error generating conversation report: {str(e)}")
            raise ExcelGenerationError(
                f"Failed to generate conversation report for session {session_id}",
                {"session_id": session_id},
                e
            )

    def generate_comparative_report(self, document_ids: List[str]) -> ExcelReport:
        """
        Generate comparative analysis report across multiple documents.
        """
        try:
            if len(document_ids) < 2:
                raise ValueError("At least 2 documents required for comparative analysis")
            
            # Get analysis data for all documents
            documents_data = []
            for doc_id in document_ids:
                document = self.document_storage.get_document(doc_id)
                analysis_data = self._extract_document_analysis_data(doc_id)
                documents_data.append({
                    'document': document,
                    'analysis': analysis_data
                })
            
            # Create comparative sheets
            sheets = [
                self._create_comparative_summary_sheet(documents_data),
                self._create_comparative_risks_sheet(documents_data),
                self._create_comparative_commitments_sheet(documents_data),
                self._create_comparative_metrics_sheet(documents_data)
            ]
            
            # Generate Excel file
            report_id = str(uuid.uuid4())
            filename = f"comparative_report_{len(document_ids)}docs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(self.reports_dir, filename)
            
            self._create_excel_file(sheets, file_path)
            
            download_url = f"/api/reports/download/{report_id}"
            
            return ExcelReport(
                report_id=report_id,
                filename=filename,
                file_path=file_path,
                download_url=download_url,
                sheets=sheets,
                created_at=datetime.now(),
                expires_at=datetime.now() + timedelta(days=7)
            )
            
        except Exception as e:
            self.logger.error(f"Error generating comparative report: {str(e)}")
            raise

    def create_custom_report(self, data_specification: Dict[str, Any]) -> ExcelReport:
        """
        Create custom report based on user specifications.
        """
        try:
            # Parse data specification
            document_ids = data_specification.get('document_ids', [])
            include_sections = data_specification.get('include_sections', ['summary'])
            custom_filters = data_specification.get('filters', {})
            
            # Collect data based on specification
            all_data = []
            for doc_id in document_ids:
                document = self.document_storage.get_document(doc_id)
                analysis_data = self._extract_document_analysis_data(doc_id)
                
                # Apply filters
                filtered_data = self._apply_custom_filters(analysis_data, custom_filters)
                all_data.append({
                    'document': document,
                    'analysis': filtered_data
                })
            
            # Create sheets based on specification
            sheets = []
            for section in include_sections:
                if section == 'summary':
                    sheets.append(self._create_custom_summary_sheet(all_data, data_specification))
                elif section == 'risks':
                    sheets.append(self._create_custom_risks_sheet(all_data, data_specification))
                elif section == 'commitments':
                    sheets.append(self._create_custom_commitments_sheet(all_data, data_specification))
                elif section == 'metrics':
                    sheets.append(self._create_custom_metrics_sheet(all_data, data_specification))
            
            # Generate Excel file
            report_id = str(uuid.uuid4())
            filename = f"custom_report_{report_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(self.reports_dir, filename)
            
            self._create_excel_file(sheets, file_path)
            
            download_url = f"/api/reports/download/{report_id}"
            
            return ExcelReport(
                report_id=report_id,
                filename=filename,
                file_path=file_path,
                download_url=download_url,
                sheets=sheets,
                created_at=datetime.now(),
                expires_at=datetime.now() + timedelta(days=7)
            )
            
        except Exception as e:
            self.logger.error(f"Error generating custom report: {str(e)}")
            raise

    def format_report_data(self, raw_data: Dict[str, Any], template: ReportTemplate) -> Dict[str, Any]:
        """
        Format raw data according to report template specifications.
        """
        formatted_data = {}
        
        for sheet_def in template.sheet_definitions:
            sheet_name = sheet_def['name']
            data_source = sheet_def['data_source']
            
            if data_source in raw_data:
                # Apply formatting rules
                formatted_data[sheet_name] = self._apply_formatting_rules(
                    raw_data[data_source],
                    template.formatting_rules.get(sheet_name, {})
                )
        
        return formatted_data

    def _extract_document_analysis_data(self, document_id: str) -> Dict[str, Any]:
        """Extract analysis data for a document (placeholder - would integrate with enhanced analyzer)."""
        # This would integrate with the enhanced summary analyzer
        # For now, return mock data structure
        return {
            'summary': {
                'document_overview': 'Sample document overview',
                'key_findings': ['Finding 1', 'Finding 2'],
                'critical_information': ['Critical info 1'],
                'recommended_actions': ['Action 1', 'Action 2'],
                'executive_recommendation': 'Sample recommendation'
            },
            'risks': [
                {
                    'risk_id': 'R001',
                    'description': 'Sample risk description',
                    'severity': 'High',
                    'category': 'Legal',
                    'affected_parties': ['Party A'],
                    'mitigation_suggestions': ['Mitigation 1'],
                    'confidence': 0.8
                }
            ],
            'commitments': [
                {
                    'commitment_id': 'C001',
                    'description': 'Sample commitment',
                    'obligated_party': 'Party A',
                    'beneficiary_party': 'Party B',
                    'deadline': '2024-12-31',
                    'status': 'Active',
                    'commitment_type': 'Deliverable'
                }
            ],
            'deliverable_dates': [
                {
                    'date': '2024-12-31',
                    'description': 'Sample deliverable',
                    'responsible_party': 'Party A',
                    'deliverable_type': 'Report',
                    'status': 'Pending'
                }
            ],
            'key_terms': ['Term 1', 'Term 2', 'Term 3']
        }

    def _extract_conversation_data(self, session_id: str) -> Dict[str, Any]:
        """Extract conversation data for a session (placeholder)."""
        # This would integrate with conversation context manager
        return {
            'session_summary': 'Sample conversation summary',
            'turns': [
                {
                    'question': 'Sample question',
                    'response': 'Sample response',
                    'timestamp': datetime.now(),
                    'analysis_mode': 'legal'
                }
            ],
            'topics': ['Topic 1', 'Topic 2']
        }

    def _create_summary_sheet(self, document: Document, analysis_data: Dict[str, Any]) -> ExcelSheet:
        """Create document summary sheet."""
        summary = analysis_data.get('summary', {})
        
        data = [
            {'Section': 'Document Name', 'Content': document.title},
            {'Section': 'Upload Date', 'Content': document.upload_timestamp.strftime('%Y-%m-%d')},
            {'Section': 'Document Overview', 'Content': summary.get('document_overview', 'N/A')},
            {'Section': 'Executive Recommendation', 'Content': summary.get('executive_recommendation', 'N/A')}
        ]
        
        # Add key findings
        for i, finding in enumerate(summary.get('key_findings', []), 1):
            data.append({'Section': f'Key Finding {i}', 'Content': finding})
        
        return ExcelSheet(
            name="Summary",
            data=data,
            formatting={'header_style': 'bold', 'wrap_text': True},
            charts=[]
        )

    def _create_risks_sheet(self, risks_data: List[Dict[str, Any]]) -> ExcelSheet:
        """Create risks analysis sheet."""
        data = []
        
        for risk in risks_data:
            data.append({
                'Risk ID': risk.get('risk_id', ''),
                'Description': risk.get('description', ''),
                'Severity': risk.get('severity', ''),
                'Category': risk.get('category', ''),
                'Affected Parties': ', '.join(risk.get('affected_parties', [])),
                'Mitigation Suggestions': ', '.join(risk.get('mitigation_suggestions', [])),
                'Confidence': risk.get('confidence', 0.0)
            })
        
        # Create chart specification for risk severity distribution
        charts = [{
            'type': 'bar',
            'title': 'Risk Distribution by Severity',
            'data_range': 'C2:C' + str(len(data) + 1),
            'categories_range': 'A2:A' + str(len(data) + 1)
        }]
        
        return ExcelSheet(
            name="Risks",
            data=data,
            formatting={'header_style': 'bold', 'conditional_formatting': {'Severity': 'risk_colors'}},
            charts=charts
        )

    def _create_commitments_sheet(self, commitments_data: List[Dict[str, Any]]) -> ExcelSheet:
        """Create commitments tracking sheet."""
        data = []
        
        for commitment in commitments_data:
            data.append({
                'Commitment ID': commitment.get('commitment_id', ''),
                'Description': commitment.get('description', ''),
                'Obligated Party': commitment.get('obligated_party', ''),
                'Beneficiary Party': commitment.get('beneficiary_party', ''),
                'Deadline': commitment.get('deadline', ''),
                'Status': commitment.get('status', ''),
                'Type': commitment.get('commitment_type', '')
            })
        
        return ExcelSheet(
            name="Commitments",
            data=data,
            formatting={'header_style': 'bold', 'date_format': 'deadline_column'},
            charts=[]
        )

    def _create_dates_sheet(self, dates_data: List[Dict[str, Any]]) -> ExcelSheet:
        """Create deliverable dates sheet."""
        data = []
        
        for date_item in dates_data:
            data.append({
                'Date': date_item.get('date', ''),
                'Description': date_item.get('description', ''),
                'Responsible Party': date_item.get('responsible_party', ''),
                'Deliverable Type': date_item.get('deliverable_type', ''),
                'Status': date_item.get('status', '')
            })
        
        return ExcelSheet(
            name="Important Dates",
            data=data,
            formatting={'header_style': 'bold', 'sort_by': 'Date'},
            charts=[]
        )

    def _create_terms_sheet(self, terms_data: List[str]) -> ExcelSheet:
        """Create key terms sheet."""
        data = [{'Term': term, 'Frequency': 1} for term in terms_data]
        
        return ExcelSheet(
            name="Key Terms",
            data=data,
            formatting={'header_style': 'bold'},
            charts=[]
        )

    def _create_conversation_summary_sheet(self, conversation_data: Dict[str, Any]) -> ExcelSheet:
        """Create conversation summary sheet."""
        data = [
            {'Metric': 'Session Summary', 'Value': conversation_data.get('session_summary', 'N/A')},
            {'Metric': 'Total Questions', 'Value': len(conversation_data.get('turns', []))},
            {'Metric': 'Topics Discussed', 'Value': ', '.join(conversation_data.get('topics', []))}
        ]
        
        return ExcelSheet(
            name="Conversation Summary",
            data=data,
            formatting={'header_style': 'bold'},
            charts=[]
        )

    def _create_qa_history_sheet(self, turns_data: List[Dict[str, Any]]) -> ExcelSheet:
        """Create Q&A history sheet."""
        data = []
        
        for i, turn in enumerate(turns_data, 1):
            data.append({
                'Turn': i,
                'Question': turn.get('question', ''),
                'Response': turn.get('response', ''),
                'Timestamp': turn.get('timestamp', '').strftime('%Y-%m-%d %H:%M:%S') if turn.get('timestamp') else '',
                'Analysis Mode': turn.get('analysis_mode', '')
            })
        
        return ExcelSheet(
            name="Q&A History",
            data=data,
            formatting={'header_style': 'bold', 'wrap_text': True},
            charts=[]
        )

    def _create_topics_sheet(self, topics_data: List[str]) -> ExcelSheet:
        """Create topics discussed sheet."""
        data = [{'Topic': topic, 'Frequency': 1} for topic in topics_data]
        
        return ExcelSheet(
            name="Topics",
            data=data,
            formatting={'header_style': 'bold'},
            charts=[]
        )

    def _create_comparative_summary_sheet(self, documents_data: List[Dict[str, Any]]) -> ExcelSheet:
        """Create comparative summary sheet."""
        data = []
        
        for doc_data in documents_data:
            document = doc_data['document']
            analysis = doc_data['analysis']
            
            data.append({
                'Document': document.title,
                'Upload Date': document.upload_timestamp.strftime('%Y-%m-%d'),
                'Risk Count': len(analysis.get('risks', [])),
                'Commitment Count': len(analysis.get('commitments', [])),
                'Key Terms Count': len(analysis.get('key_terms', []))
            })
        
        return ExcelSheet(
            name="Comparative Summary",
            data=data,
            formatting={'header_style': 'bold'},
            charts=[{
                'type': 'bar',
                'title': 'Document Comparison',
                'data_range': 'C2:E' + str(len(data) + 1),
                'categories_range': 'A2:A' + str(len(data) + 1)
            }]
        )

    def _create_comparative_risks_sheet(self, documents_data: List[Dict[str, Any]]) -> ExcelSheet:
        """Create comparative risks sheet."""
        data = []
        
        for doc_data in documents_data:
            document = doc_data['document']
            risks = doc_data['analysis'].get('risks', [])
            
            for risk in risks:
                data.append({
                    'Document': document.title,
                    'Risk ID': risk.get('risk_id', ''),
                    'Description': risk.get('description', ''),
                    'Severity': risk.get('severity', ''),
                    'Category': risk.get('category', '')
                })
        
        return ExcelSheet(
            name="All Risks",
            data=data,
            formatting={'header_style': 'bold'},
            charts=[]
        )

    def _create_comparative_commitments_sheet(self, documents_data: List[Dict[str, Any]]) -> ExcelSheet:
        """Create comparative commitments sheet."""
        data = []
        
        for doc_data in documents_data:
            document = doc_data['document']
            commitments = doc_data['analysis'].get('commitments', [])
            
            for commitment in commitments:
                data.append({
                    'Document': document.title,
                    'Commitment ID': commitment.get('commitment_id', ''),
                    'Description': commitment.get('description', ''),
                    'Obligated Party': commitment.get('obligated_party', ''),
                    'Deadline': commitment.get('deadline', ''),
                    'Status': commitment.get('status', '')
                })
        
        return ExcelSheet(
            name="All Commitments",
            data=data,
            formatting={'header_style': 'bold'},
            charts=[]
        )

    def _create_comparative_metrics_sheet(self, documents_data: List[Dict[str, Any]]) -> ExcelSheet:
        """Create comparative metrics sheet."""
        data = []
        
        for doc_data in documents_data:
            document = doc_data['document']
            analysis = doc_data['analysis']
            risks = analysis.get('risks', [])
            
            # Calculate metrics
            high_risks = len([r for r in risks if r.get('severity') == 'High'])
            medium_risks = len([r for r in risks if r.get('severity') == 'Medium'])
            low_risks = len([r for r in risks if r.get('severity') == 'Low'])
            
            data.append({
                'Document': document.title,
                'Total Risks': len(risks),
                'High Risk Count': high_risks,
                'Medium Risk Count': medium_risks,
                'Low Risk Count': low_risks,
                'Risk Score': (high_risks * 3) + (medium_risks * 2) + (low_risks * 1)
            })
        
        return ExcelSheet(
            name="Risk Metrics",
            data=data,
            formatting={'header_style': 'bold'},
            charts=[{
                'type': 'bar',
                'title': 'Risk Scores by Document',
                'data_range': 'F2:F' + str(len(data) + 1),
                'categories_range': 'A2:A' + str(len(data) + 1)
            }]
        )

    def _create_custom_summary_sheet(self, all_data: List[Dict[str, Any]], specification: Dict[str, Any]) -> ExcelSheet:
        """Create custom summary sheet based on specification."""
        # Implementation would depend on specific custom requirements
        return self._create_comparative_summary_sheet(all_data)

    def _create_custom_risks_sheet(self, all_data: List[Dict[str, Any]], specification: Dict[str, Any]) -> ExcelSheet:
        """Create custom risks sheet based on specification."""
        return self._create_comparative_risks_sheet(all_data)

    def _create_custom_commitments_sheet(self, all_data: List[Dict[str, Any]], specification: Dict[str, Any]) -> ExcelSheet:
        """Create custom commitments sheet based on specification."""
        return self._create_comparative_commitments_sheet(all_data)

    def _create_custom_metrics_sheet(self, all_data: List[Dict[str, Any]], specification: Dict[str, Any]) -> ExcelSheet:
        """Create custom metrics sheet based on specification."""
        return self._create_comparative_metrics_sheet(all_data)

    def _apply_custom_filters(self, analysis_data: Dict[str, Any], filters: Dict[str, Any]) -> Dict[str, Any]:
        """Apply custom filters to analysis data."""
        filtered_data = analysis_data.copy()
        
        # Apply severity filter for risks
        if 'risk_severity' in filters:
            allowed_severities = filters['risk_severity']
            filtered_data['risks'] = [
                risk for risk in filtered_data.get('risks', [])
                if risk.get('severity') in allowed_severities
            ]
        
        # Apply date range filter for commitments
        if 'date_range' in filters:
            start_date = filters['date_range'].get('start')
            end_date = filters['date_range'].get('end')
            # Filter logic would be implemented here
        
        return filtered_data

    def _apply_formatting_rules(self, data: Any, formatting_rules: Dict[str, Any]) -> Any:
        """Apply formatting rules to data."""
        # Implementation would depend on specific formatting requirements
        return data

    def _generate_fallback_report(self, document_id: str, analysis_data: Dict[str, Any], 
                                 report_type: str) -> ExcelReport:
        """Generate fallback report in alternative format when Excel generation fails."""
        self.logger.info(f"Generating fallback report for document {document_id}")
        
        # Create simplified data structure
        fallback_data = {
            'document_summary': analysis_data.get('summary', {}),
            'risks': analysis_data.get('risks', []),
            'commitments': analysis_data.get('commitments', []),
            'key_dates': analysis_data.get('deliverable_dates', [])
        }
        
        # Generate CSV format as fallback
        csv_content = alternative_formats.create_fallback_report(fallback_data, "csv")
        
        # Save as text file
        report_id = str(uuid.uuid4())
        filename = f"fallback_report_{document_id}_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        file_path = os.path.join(self.reports_dir, filename)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(csv_content)
        
        # Create minimal ExcelReport object
        return ExcelReport(
            report_id=report_id,
            filename=filename,
            file_path=file_path,
            download_url=f"/api/reports/download/{report_id}",
            sheets=[ExcelSheet(
                name="Fallback Report",
                data=[{"note": "Excel generation failed, CSV format provided"}],
                formatting={},
                charts=[]
            )],
            created_at=datetime.now(),
            expires_at=datetime.now() + timedelta(days=7)
        )
    
    def _create_excel_file(self, sheets: List[ExcelSheet], file_path: str) -> None:
        """Create the actual Excel file with all sheets and formatting."""
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        for sheet_data in sheets:
            worksheet = workbook.create_sheet(title=sheet_data.name)
            
            if sheet_data.data:
                # Convert data to DataFrame for easier handling
                df = pd.DataFrame(sheet_data.data)
                
                # Write data to worksheet
                for r in dataframe_to_rows(df, index=False, header=True):
                    worksheet.append(r)
                
                # Apply formatting
                self._apply_excel_formatting(worksheet, sheet_data.formatting)
                
                # Add charts if specified
                for chart_spec in sheet_data.charts:
                    self._add_chart_to_worksheet(worksheet, chart_spec)
        
        # Save workbook with error handling
        try:
            workbook.save(file_path)
        except Exception as e:
            self.logger.error(f"Failed to save Excel file: {str(e)}")
            raise ExcelGenerationError(
                f"Failed to save Excel file to {file_path}",
                {"file_path": file_path},
                e
            )

    def _apply_excel_formatting(self, worksheet, formatting: Dict[str, Any]) -> None:
        """Apply formatting to Excel worksheet."""
        if not formatting:
            return
        
        # Apply header formatting
        if formatting.get('header_style') == 'bold':
            for cell in worksheet[1]:  # First row
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal="center")
        
        # Apply text wrapping
        if formatting.get('wrap_text'):
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = self.data_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _add_chart_to_worksheet(self, worksheet, chart_spec: Dict[str, Any]) -> None:
        """Add chart to worksheet based on specification."""
        if chart_spec.get('type') == 'bar':
            chart = BarChart()
            chart.title = chart_spec.get('title', 'Chart')
            
            # This would need proper range handling based on actual data
            # For now, just create a basic chart structure
            chart.x_axis.title = 'Categories'
            chart.y_axis.title = 'Values'
            
            # Add chart to worksheet (position would be calculated)
            worksheet.add_chart(chart, "H2")