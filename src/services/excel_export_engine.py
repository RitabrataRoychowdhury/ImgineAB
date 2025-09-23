"""
Enhanced Excel Export Engine with Never-Fail Guarantee
Provides guaranteed export generation with multiple fallback mechanisms.
"""

import os
import uuid
import json
import csv
import logging
from typing import List, Dict, Any, Optional, Union, Tuple
from datetime import datetime, timedelta
from dataclasses import dataclass, asdict
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.utils.logging_config import get_logger
from src.storage.document_storage import DocumentStorage

logger = get_logger(__name__)


@dataclass
class ExportResult:
    """Result of export generation with metadata."""
    success: bool
    file_path: str
    filename: str
    download_url: str
    format_type: str  # 'excel', 'csv', 'json', 'text'
    expiration_time: datetime
    metadata: Dict[str, Any]
    error_message: Optional[str] = None
    fallback_used: bool = False


@dataclass
class ExcelSheet:
    """Excel sheet configuration."""
    name: str
    data: List[Dict[str, Any]]
    formatting: Dict[str, Any]
    charts: List[Dict[str, Any]]
    pivot_tables: List[Dict[str, Any]] = None
    conditional_formatting: List[Dict[str, Any]] = None


@dataclass
class ExportRequest:
    """Export request specification."""
    data: Any
    format_preferences: List[str]  # ['excel', 'csv', 'json']
    template_type: str  # 'risk_analysis', 'document_summary', 'portfolio'
    metadata: Dict[str, Any]
    user_request: str


class ExcelExportEngine:
    """
    Never-fail Excel export engine with comprehensive fallback mechanisms.
    Guarantees that user requests for tabular data always result in downloadable files.
    """
    
    def __init__(self, storage: DocumentStorage, exports_dir: str = "data/exports"):
        self.storage = storage
        self.exports_dir = Path(exports_dir)
        self.exports_dir.mkdir(parents=True, exist_ok=True)
        
        # Cleanup old files on initialization
        self._cleanup_expired_files()
        
        # Default styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.data_alignment = Alignment(wrap_text=True, vertical="top")
        
        # Export templates
        self.templates = {
            'risk_analysis': self._get_risk_analysis_template,
            'document_summary': self._get_document_summary_template,
            'portfolio_analysis': self._get_portfolio_analysis_template,
            'qa_history': self._get_qa_history_template,
            'comparative_analysis': self._get_comparative_analysis_template
        }
    
    def export_data(self, request: ExportRequest) -> ExportResult:
        """
        Main export method with never-fail guarantee.
        Always returns a downloadable file, using fallbacks if necessary.
        """
        logger.info(f"Starting export for request: {request.template_type}")
        
        # Try primary export path
        try:
            return self._try_primary_export(request)
        except Exception as e1:
            logger.warning(f"Primary export failed: {e1}")
            
            # Try fallback 1: Basic Excel
            try:
                return self._try_basic_excel_export(request)
            except Exception as e2:
                logger.warning(f"Basic Excel export failed: {e2}")
                
                # Try fallback 2: CSV export
                try:
                    return self._try_csv_export(request)
                except Exception as e3:
                    logger.warning(f"CSV export failed: {e3}")
                    
                    # Final fallback: Structured text
                    return self._generate_text_export(request)
    
    def generate_risk_analysis_report(self, risk_data: Dict[str, Any], 
                                    document_info: Dict[str, Any]) -> ExportResult:
        """Generate comprehensive risk analysis Excel report."""
        request = ExportRequest(
            data={'risks': risk_data, 'document': document_info},
            format_preferences=['excel', 'csv'],
            template_type='risk_analysis',
            metadata={'report_type': 'risk_analysis'},
            user_request='Generate risk analysis report'
        )
        return self.export_data(request)
    
    def generate_document_summary_export(self, summary_data: Dict[str, Any]) -> ExportResult:
        """Generate document summary Excel report."""
        request = ExportRequest(
            data=summary_data,
            format_preferences=['excel', 'csv'],
            template_type='document_summary',
            metadata={'report_type': 'document_summary'},
            user_request='Generate document summary export'
        )
        return self.export_data(request)
    
    def generate_portfolio_analysis_report(self, portfolio_data: List[Dict[str, Any]]) -> ExportResult:
        """Generate portfolio analysis Excel report."""
        request = ExportRequest(
            data={'documents': portfolio_data},
            format_preferences=['excel', 'csv'],
            template_type='portfolio_analysis',
            metadata={'report_type': 'portfolio_analysis'},
            user_request='Generate portfolio analysis report'
        )
        return self.export_data(request)
    
    def export_tabular_data(self, data: Any, user_request: str, 
                          format_preferences: List[str] = None) -> ExportResult:
        """Export any tabular data with automatic format detection."""
        if format_preferences is None:
            format_preferences = ['excel', 'csv', 'json']
        
        # Detect data structure and choose appropriate template
        template_type = self._detect_template_type(data, user_request)
        
        request = ExportRequest(
            data=data,
            format_preferences=format_preferences,
            template_type=template_type,
            metadata={'auto_detected': True},
            user_request=user_request
        )
        return self.export_data(request)
    
    def _try_primary_export(self, request: ExportRequest) -> ExportResult:
        """Try primary export with full Excel formatting and features."""
        logger.info("Attempting primary Excel export with full formatting")
        
        # Get template
        template_func = self.templates.get(request.template_type, self._get_generic_template)
        sheets = template_func(request.data)
        
        # Generate Excel file
        filename = self._generate_filename(request.template_type, 'xlsx')
        file_path = self.exports_dir / filename
        
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet
        
        for sheet_config in sheets:
            self._create_excel_sheet(workbook, sheet_config)
        
        workbook.save(str(file_path))
        
        return ExportResult(
            success=True,
            file_path=str(file_path),
            filename=filename,
            download_url=self._generate_download_url(filename),
            format_type='excel',
            expiration_time=datetime.now() + timedelta(hours=24),
            metadata=request.metadata,
            fallback_used=False
        )
    
    def _try_basic_excel_export(self, request: ExportRequest) -> ExportResult:
        """Try basic Excel export without advanced formatting."""
        logger.info("Attempting basic Excel export")
        
        filename = self._generate_filename(request.template_type, 'xlsx')
        file_path = self.exports_dir / filename
        
        # Create basic Excel with pandas
        data_frames = self._convert_to_dataframes(request.data)
        
        with pd.ExcelWriter(str(file_path), engine='openpyxl') as writer:
            for sheet_name, df in data_frames.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        return ExportResult(
            success=True,
            file_path=str(file_path),
            filename=filename,
            download_url=self._generate_download_url(filename),
            format_type='excel',
            expiration_time=datetime.now() + timedelta(hours=24),
            metadata=request.metadata,
            fallback_used=True
        )
    
    def _try_csv_export(self, request: ExportRequest) -> ExportResult:
        """Try CSV export as fallback."""
        logger.info("Attempting CSV export")
        
        filename = self._generate_filename(request.template_type, 'csv')
        file_path = self.exports_dir / filename
        
        # Convert data to single CSV
        main_data = self._extract_main_data_for_csv(request.data)
        
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            if main_data:
                fieldnames = main_data[0].keys()
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(main_data)
        
        return ExportResult(
            success=True,
            file_path=str(file_path),
            filename=filename,
            download_url=self._generate_download_url(filename),
            format_type='csv',
            expiration_time=datetime.now() + timedelta(hours=24),
            metadata=request.metadata,
            fallback_used=True
        )
    
    def _generate_text_export(self, request: ExportRequest) -> ExportResult:
        """Final fallback: Generate structured text export."""
        logger.info("Generating structured text export as final fallback")
        
        filename = self._generate_filename(request.template_type, 'txt')
        file_path = self.exports_dir / filename
        
        # Create structured text representation
        text_content = self._format_data_as_text(request.data, request.user_request)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(text_content)
        
        return ExportResult(
            success=True,
            file_path=str(file_path),
            filename=filename,
            download_url=self._generate_download_url(filename),
            format_type='text',
            expiration_time=datetime.now() + timedelta(hours=24),
            metadata=request.metadata,
            fallback_used=True
        )
    
    def _get_risk_analysis_template(self, data: Dict[str, Any]) -> List[ExcelSheet]:
        """Generate risk analysis template sheets."""
        risks = data.get('risks', [])
        document_info = data.get('document', {})
        
        sheets = []
        
        # Executive Summary Sheet
        summary_data = [
            {'Metric': 'Document Name', 'Value': document_info.get('title', 'Unknown')},
            {'Metric': 'Analysis Date', 'Value': datetime.now().strftime('%Y-%m-%d')},
            {'Metric': 'Total Risks Identified', 'Value': len(risks)},
            {'Metric': 'High Risk Count', 'Value': len([r for r in risks if r.get('severity') == 'High'])},
            {'Metric': 'Medium Risk Count', 'Value': len([r for r in risks if r.get('severity') == 'Medium'])},
            {'Metric': 'Low Risk Count', 'Value': len([r for r in risks if r.get('severity') == 'Low'])},
        ]
        
        sheets.append(ExcelSheet(
            name="Executive Summary",
            data=summary_data,
            formatting={'header_style': True, 'wrap_text': True},
            charts=[{
                'type': 'pie',
                'title': 'Risk Distribution by Severity',
                'data_range': 'B4:B6',
                'categories_range': 'A4:A6'
            }]
        ))
        
        # Risk Matrix Sheet
        risk_matrix_data = []
        for risk in risks:
            risk_matrix_data.append({
                'Risk ID': risk.get('risk_id', f"R{len(risk_matrix_data)+1:03d}"),
                'Description': risk.get('description', ''),
                'Category': risk.get('category', 'General'),
                'Severity': risk.get('severity', 'Medium'),
                'Probability': risk.get('probability', 0.5),
                'Impact': risk.get('impact_description', ''),
                'Affected Parties': ', '.join(risk.get('affected_parties', [])),
                'Mitigation Strategies': ', '.join(risk.get('mitigation_strategies', [])),
                'Timeline': risk.get('materialization_timeframe', 'Unknown')
            })
        
        sheets.append(ExcelSheet(
            name="Risk Matrix",
            data=risk_matrix_data,
            formatting={'header_style': True, 'conditional_formatting': True},
            charts=[{
                'type': 'bar',
                'title': 'Risks by Category',
                'data_range': 'C2:C' + str(len(risk_matrix_data) + 1),
                'categories_range': 'A2:A' + str(len(risk_matrix_data) + 1)
            }]
        ))
        
        # Action Plan Sheet
        action_data = []
        for i, risk in enumerate(risks, 1):
            for j, mitigation in enumerate(risk.get('mitigation_strategies', []), 1):
                action_data.append({
                    'Action ID': f"A{i:03d}-{j}",
                    'Related Risk': risk.get('risk_id', f"R{i:03d}"),
                    'Action Description': mitigation,
                    'Priority': risk.get('severity', 'Medium'),
                    'Responsible Party': risk.get('affected_parties', ['TBD'])[0] if risk.get('affected_parties') else 'TBD',
                    'Target Date': 'TBD',
                    'Status': 'Pending'
                })
        
        sheets.append(ExcelSheet(
            name="Action Plan",
            data=action_data,
            formatting={'header_style': True, 'date_format': True},
            charts=[]
        ))
        
        return sheets
    
    def _get_document_summary_template(self, data: Dict[str, Any]) -> List[ExcelSheet]:
        """Generate document summary template sheets."""
        sheets = []
        
        # Overview Sheet
        overview_data = [
            {'Section': 'Document Overview', 'Content': data.get('overview', 'N/A')},
            {'Section': 'Document Type', 'Content': data.get('document_type', 'Unknown')},
            {'Section': 'Analysis Date', 'Content': datetime.now().strftime('%Y-%m-%d')},
            {'Section': 'Key Findings Count', 'Content': len(data.get('key_terms', []))},
        ]
        
        sheets.append(ExcelSheet(
            name="Overview",
            data=overview_data,
            formatting={'header_style': True, 'wrap_text': True},
            charts=[]
        ))
        
        # Key Terms Sheet
        key_terms = data.get('key_terms', {})
        terms_data = [{'Term': term, 'Definition': definition} 
                     for term, definition in key_terms.items()]
        
        sheets.append(ExcelSheet(
            name="Key Terms",
            data=terms_data,
            formatting={'header_style': True, 'wrap_text': True},
            charts=[]
        ))
        
        # Important Dates Sheet
        dates = data.get('important_dates', [])
        dates_data = []
        for date_info in dates:
            dates_data.append({
                'Date': date_info.get('date', ''),
                'Description': date_info.get('description', ''),
                'Type': date_info.get('type', 'General'),
                'Responsible Party': date_info.get('responsible_party', 'N/A')
            })
        
        sheets.append(ExcelSheet(
            name="Important Dates",
            data=dates_data,
            formatting={'header_style': True, 'date_format': True},
            charts=[]
        ))
        
        return sheets
    
    def _get_portfolio_analysis_template(self, data: Dict[str, Any]) -> List[ExcelSheet]:
        """Generate portfolio analysis template sheets."""
        documents = data.get('documents', [])
        sheets = []
        
        # Portfolio Summary Sheet
        summary_data = [
            {'Metric': 'Total Documents', 'Value': len(documents)},
            {'Metric': 'Analysis Date', 'Value': datetime.now().strftime('%Y-%m-%d')},
            {'Metric': 'Document Types', 'Value': len(set(doc.get('document_type', 'Unknown') for doc in documents))},
        ]
        
        # Add document type breakdown
        doc_types = {}
        for doc in documents:
            doc_type = doc.get('document_type', 'Unknown')
            doc_types[doc_type] = doc_types.get(doc_type, 0) + 1
        
        for doc_type, count in doc_types.items():
            summary_data.append({'Metric': f'{doc_type} Count', 'Value': count})
        
        sheets.append(ExcelSheet(
            name="Portfolio Summary",
            data=summary_data,
            formatting={'header_style': True},
            charts=[{
                'type': 'pie',
                'title': 'Document Types Distribution',
                'data_range': f'B4:B{3 + len(doc_types)}',
                'categories_range': f'A4:A{3 + len(doc_types)}'
            }]
        ))
        
        # Document List Sheet
        doc_list_data = []
        for doc in documents:
            doc_list_data.append({
                'Document Name': doc.get('title', 'Unknown'),
                'Type': doc.get('document_type', 'Unknown'),
                'Upload Date': doc.get('upload_date', ''),
                'Status': doc.get('status', 'Unknown'),
                'Risk Count': len(doc.get('risks', [])),
                'Key Terms Count': len(doc.get('key_terms', []))
            })
        
        sheets.append(ExcelSheet(
            name="Document List",
            data=doc_list_data,
            formatting={'header_style': True, 'date_format': True},
            charts=[]
        ))
        
        return sheets
    
    def _get_qa_history_template(self, data: Dict[str, Any]) -> List[ExcelSheet]:
        """Generate Q&A history template sheets."""
        qa_history = data.get('qa_history', [])
        
        sheets = []
        
        # Q&A History Sheet
        qa_data = []
        for i, qa in enumerate(qa_history, 1):
            qa_data.append({
                'Question #': i,
                'Question': qa.get('question', ''),
                'Answer': qa.get('answer', ''),
                'Timestamp': qa.get('timestamp', ''),
                'Analysis Mode': qa.get('analysis_mode', 'Standard'),
                'Confidence': qa.get('confidence', 'N/A')
            })
        
        sheets.append(ExcelSheet(
            name="Q&A History",
            data=qa_data,
            formatting={'header_style': True, 'wrap_text': True},
            charts=[]
        ))
        
        return sheets
    
    def _get_comparative_analysis_template(self, data: Dict[str, Any]) -> List[ExcelSheet]:
        """Generate comparative analysis template sheets."""
        # Implementation for comparative analysis
        return self._get_portfolio_analysis_template(data)
    
    def _get_generic_template(self, data: Any) -> List[ExcelSheet]:
        """Generate generic template for unknown data types."""
        sheets = []
        
        # Try to convert data to tabular format
        if isinstance(data, dict):
            # Convert dict to key-value pairs
            table_data = [{'Key': k, 'Value': str(v)} for k, v in data.items()]
        elif isinstance(data, list) and data and isinstance(data[0], dict):
            # Already in tabular format
            table_data = data
        else:
            # Convert to string representation
            table_data = [{'Data': str(data)}]
        
        sheets.append(ExcelSheet(
            name="Data",
            data=table_data,
            formatting={'header_style': True, 'wrap_text': True},
            charts=[]
        ))
        
        return sheets
    
    def _create_excel_sheet(self, workbook: Workbook, sheet_config: ExcelSheet):
        """Create an Excel sheet with formatting and charts."""
        worksheet = workbook.create_sheet(title=sheet_config.name)
        
        if not sheet_config.data:
            return
        
        # Convert to DataFrame for easier handling
        df = pd.DataFrame(sheet_config.data)
        
        # Write data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(r)
        
        # Apply formatting
        if sheet_config.formatting.get('header_style'):
            self._apply_header_formatting(worksheet)
        
        if sheet_config.formatting.get('wrap_text'):
            self._apply_wrap_text(worksheet)
        
        if sheet_config.formatting.get('conditional_formatting'):
            self._apply_conditional_formatting(worksheet, df)
        
        if sheet_config.formatting.get('date_format'):
            self._apply_date_formatting(worksheet, df)
        
        # Add charts
        for chart_config in sheet_config.charts:
            self._add_chart(worksheet, chart_config, len(df))
        
        # Auto-adjust column widths
        self._auto_adjust_columns(worksheet)
    
    def _apply_header_formatting(self, worksheet):
        """Apply header formatting to the first row."""
        for cell in worksheet[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _apply_wrap_text(self, worksheet):
        """Apply wrap text to all cells."""
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = self.data_alignment
    
    def _apply_conditional_formatting(self, worksheet, df):
        """Apply conditional formatting based on data values."""
        # Example: Color-code severity levels
        if 'Severity' in df.columns:
            severity_col = df.columns.get_loc('Severity') + 1  # Excel is 1-indexed
            
            # High severity - red
            worksheet.conditional_formatting.add(
                f'{worksheet.cell(2, severity_col).coordinate}:{worksheet.cell(len(df)+1, severity_col).coordinate}',
                {'type': 'containsText', 'operator': 'containsText', 'text': 'High', 'format': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")}
            )
    
    def _apply_date_formatting(self, worksheet, df):
        """Apply date formatting to date columns."""
        date_columns = [col for col in df.columns if 'date' in col.lower() or 'Date' in col]
        
        for col_name in date_columns:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                for row in range(2, len(df) + 2):  # Skip header
                    cell = worksheet.cell(row, col_idx)
                    cell.number_format = 'YYYY-MM-DD'
    
    def _add_chart(self, worksheet, chart_config: Dict[str, Any], data_rows: int):
        """Add chart to worksheet."""
        chart_type = chart_config.get('type', 'bar')
        
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'pie':
            chart = PieChart()
        elif chart_type == 'line':
            chart = LineChart()
        else:
            return  # Unsupported chart type
        
        chart.title = chart_config.get('title', 'Chart')
        
        # Set data references
        data_range = chart_config.get('data_range', f'B2:B{data_rows + 1}')
        categories_range = chart_config.get('categories_range', f'A2:A{data_rows + 1}')
        
        data = Reference(worksheet, range_string=data_range)
        categories = Reference(worksheet, range_string=categories_range)
        
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(categories)
        
        # Position chart
        chart_position = chart_config.get('position', 'E2')
        worksheet.add_chart(chart, chart_position)
    
    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths based on content."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _convert_to_dataframes(self, data: Any) -> Dict[str, pd.DataFrame]:
        """Convert data to pandas DataFrames for basic Excel export."""
        dataframes = {}
        
        if isinstance(data, dict):
            for key, value in data.items():
                if isinstance(value, list) and value and isinstance(value[0], dict):
                    dataframes[key] = pd.DataFrame(value)
                elif isinstance(value, dict):
                    dataframes[key] = pd.DataFrame([value])
                else:
                    dataframes[key] = pd.DataFrame([{key: value}])
        elif isinstance(data, list) and data and isinstance(data[0], dict):
            dataframes['Data'] = pd.DataFrame(data)
        else:
            dataframes['Data'] = pd.DataFrame([{'Value': str(data)}])
        
        return dataframes
    
    def _extract_main_data_for_csv(self, data: Any) -> List[Dict[str, Any]]:
        """Extract main data for CSV export."""
        if isinstance(data, list) and data and isinstance(data[0], dict):
            return data
        elif isinstance(data, dict):
            # Find the largest list of dicts in the data
            largest_list = []
            for value in data.values():
                if isinstance(value, list) and value and isinstance(value[0], dict):
                    if len(value) > len(largest_list):
                        largest_list = value
            
            if largest_list:
                return largest_list
            else:
                # Convert dict to list of key-value pairs
                return [{'Key': k, 'Value': str(v)} for k, v in data.items()]
        else:
            return [{'Data': str(data)}]
    
    def _format_data_as_text(self, data: Any, user_request: str) -> str:
        """Format data as structured text."""
        lines = [
            f"Data Export - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"User Request: {user_request}",
            "=" * 50,
            ""
        ]
        
        if isinstance(data, dict):
            for key, value in data.items():
                lines.append(f"{key.upper()}:")
                lines.append("-" * len(key))
                
                if isinstance(value, list):
                    for i, item in enumerate(value, 1):
                        if isinstance(item, dict):
                            lines.append(f"{i}. {json.dumps(item, indent=2)}")
                        else:
                            lines.append(f"{i}. {item}")
                else:
                    lines.append(str(value))
                
                lines.append("")
        else:
            lines.append("DATA:")
            lines.append(str(data))
        
        return "\n".join(lines)
    
    def _detect_template_type(self, data: Any, user_request: str) -> str:
        """Detect appropriate template type based on data and request."""
        request_lower = user_request.lower()
        
        if 'risk' in request_lower:
            return 'risk_analysis'
        elif 'summary' in request_lower:
            return 'document_summary'
        elif 'portfolio' in request_lower or 'multiple' in request_lower:
            return 'portfolio_analysis'
        elif 'question' in request_lower or 'qa' in request_lower:
            return 'qa_history'
        elif 'compare' in request_lower or 'comparison' in request_lower:
            return 'comparative_analysis'
        else:
            return 'generic'
    
    def _generate_filename(self, template_type: str, extension: str) -> str:
        """Generate unique filename for export."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_id = str(uuid.uuid4())[:8]
        return f"{template_type}_{timestamp}_{unique_id}.{extension}"
    
    def _generate_download_url(self, filename: str) -> str:
        """Generate download URL for the exported file."""
        return f"/api/exports/download/{filename}"
    
    def _cleanup_expired_files(self):
        """Clean up expired export files."""
        try:
            current_time = datetime.now()
            for file_path in self.exports_dir.glob("*"):
                if file_path.is_file():
                    # Check if file is older than 24 hours
                    file_time = datetime.fromtimestamp(file_path.stat().st_mtime)
                    if current_time - file_time > timedelta(hours=24):
                        file_path.unlink()
                        logger.info(f"Cleaned up expired file: {file_path.name}")
        except Exception as e:
            logger.warning(f"Error during file cleanup: {e}")
    
    def get_download_info(self, filename: str) -> Optional[Dict[str, Any]]:
        """Get download information for a file."""
        file_path = self.exports_dir / filename
        if file_path.exists():
            return {
                'filename': filename,
                'file_path': str(file_path),
                'size': file_path.stat().st_size,
                'created': datetime.fromtimestamp(file_path.stat().st_ctime),
                'download_url': self._generate_download_url(filename)
            }
        return None